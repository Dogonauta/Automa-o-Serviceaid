import openpyxl
from dotenv import load_dotenv
import os
from selenium import webdriver
from selenium.webdriver.chrome.service import Service
from selenium.webdriver.common.by import By
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from time import sleep

possiveis_status = [
    "Em andamento", "Resolvido", "Execução pendente",
    "Resposta do Cliente", "Pendente Material", "Aceitação",
    "Cadastro contábil", "Pendente Fornecedor"
]
load_dotenv()

User = os.getenv("User")
Senha = os.getenv("Senha")

lista_chamados = openpyxl.load_workbook('tickets.xlsx')
pagina_lista_chamados = lista_chamados['Plan1']

driver = webdriver.Chrome()
driver.get('https://csm2.serviceaide.com/#login')
sleep(2)
def logar():

    login = WebDriverWait(driver, 20).until(
    EC.presence_of_element_located((By.XPATH, "//input[@placeholder='USER NAME']"))
    )
    login.send_keys(User)
    login.send_keys(Keys.TAB)

    senha = driver.find_element(By.XPATH, "//input[@name='password']")
    senha.send_keys(Senha)

    confirmar = WebDriverWait(driver, 20).until(
    EC.presence_of_element_located((By.XPATH, "//span[contains(text(),'SIGN IN')]"))
    )
    confirmar.click()
    sleep(0.5)

def testar_status():
    status_chamado = WebDriverWait(driver, 20).until(
    EC.presence_of_element_located((By.XPATH, "/html/body/div[3]/div[2]/div[1]/div/div[2]/div[3]/div/div[2]/div/div/div/div[2]/div[2]/div[1]/div[2]/div[2]/div[6]/div/div/div[3]"))
    )
    sleep(1)
    if status_chamado.text not in possiveis_status:
        status_chamado = WebDriverWait(driver, 20).until(
        EC.presence_of_element_located((By.XPATH, "/html/body/div[3]/div[2]/div[1]/div/div[2]/div[3]/div/div[2]/div/div/div/div[2]/div[2]/div[1]/div[2]/div[2]/div[7]/div/div/div[3]"))
        )
    
    return status_chamado

def pesquisar_chamado():
    pesquisa = WebDriverWait(driver, 20).until(
    EC.presence_of_element_located((By.XPATH, "//span[contains(text(),'Pesquisar')]"))
    )
    pesquisa.click()
    sleep(0.5)

    consulta = WebDriverWait(driver, 20).until(
    EC.presence_of_element_located((By.XPATH, "//input[@autoidentifier='txtbox_GSSearchServiceDesk']"))
    )
    consulta.clear()
    consulta.send_keys(chamado)
    consulta.send_keys(Keys.ENTER)
        
    caso = WebDriverWait(driver, 50).until(
    EC.presence_of_element_located((By.XPATH, "//a[@tabindex='1']"))
    )
    caso.click()

def atender_chamado():
    sleep(0.5)
    apontamento = WebDriverWait(driver, 20).until(
    EC.presence_of_element_located((By.XPATH, "//span[contains(text(),'Apontamentos') and contains(@class, 'x-tab-inner')]"))
    )
    apontamento.click()
    sleep(0.5)

    combo_box_apontamento = WebDriverWait(driver, 20).until(
    EC.presence_of_element_located((By.XPATH, "//input[@autoidentifier='combo_WorklogType']"))
    )
    combo_box_apontamento.click()
    sleep(0.5)

    tipo_apontamento = WebDriverWait(driver, 20).until(
    EC.presence_of_element_located((By.XPATH, "//li[contains(text(),'Atualizar status')]"))
    )
    tipo_apontamento.click()
    sleep(0.5)

    texto_apontamento = WebDriverWait(driver, 20).until(
    EC.presence_of_element_located((By.XPATH, "//textarea[@autoidentifier=('txtbox_Worklog')]"))
    )
    texto_apontamento.send_keys("Chamado encerrado conforme inventário 2023.")
    sleep(0.5)

    visivel_cliente = WebDriverWait(driver, 20).until(
    EC.presence_of_element_located((By.XPATH, "//input[@autoidentifier=('chkbox_ClientViewable')]"))
    )
    visivel_cliente.click()
    sleep(0.5)
    botao_de_acao()
    sleep(0.5)

    try: 
        botao_atualizar_cadastro()
        sleep(1)
    except:
        print("chamado sem botao de atualizar cadastro")
    resolver_chamado()

def resolver_chamado():
    sleep(1)
    botao_de_acao()
    sleep(1)
    try:
        botao_de_retomar()
    except:
        print("chamado ja aberto, e sem botão de 'retomar'")
    sleep(1)
    dar_resolucao()
    sleep(1.5)
    botao_de_acao()
    sleep(0.5)
    botao_de_resolver()
    sleep(2)
    alimentar_nova_planilha()
    fechar_chamado()
    
def fechar_chamado():
    sleep(0.5)
    subir_aba()
    fechar_aba = WebDriverWait(driver, 2).until(
    EC.element_to_be_clickable((By.XPATH, "//a[@class='x-tab-close-btn']"))
    )
    fechar_aba.click()

def botao_de_acao():
    acao = WebDriverWait(driver, 20).until(
    EC.visibility_of_element_located((By.XPATH, "//button[@id='ca-global-actions-menu-btnEl']"))
    )
    WebDriverWait(driver, 20).until(
        EC.element_to_be_clickable((acao))
    )
    acao.click()

def botao_de_aceitacao():
    aceitar_solicitacao = WebDriverWait(driver, 20).until(
    EC.presence_of_element_located((By.XPATH, "//span[contains(text(),'Aceitar solicitação')]"))
    )
    aceitar_solicitacao.click()

def botao_atualizar_cadastro():
    atualizar_cadastro = WebDriverWait(driver, 2).until(
    EC.presence_of_element_located((By.XPATH, "//span[contains(text(), 'Atualizar cadastro')]"))
    )
    atualizar_cadastro.click()

def botao_de_retomar():
    retomar = WebDriverWait(driver, 5).until(
    EC.presence_of_element_located((By.XPATH, "//span[contains(text(),'Retomar solicitação')]"))
    )
    retomar.click()

def botao_de_resolver():
    resolver = WebDriverWait(driver, 20).until(
    EC.presence_of_element_located((By.XPATH, "//span[contains(text(),'Resolver solicitação')]"))
    )
    resolver.click()

def dar_resolucao():
    resolucao = WebDriverWait(driver, 20).until(
    EC.presence_of_element_located((By.XPATH, "//textarea[@autoidentifier='txtbox_Resolution']"))
    ) 
    resolucao.send_keys(Keys.CONTROL + "a") 
    resolucao.send_keys("Chamado encerrado conforme inventário 2023.")
    sleep(1)
    resolucao.send_keys(Keys.TAB)

def alimentar_nova_planilha():
    resolucao = WebDriverWait(driver, 20).until(
    EC.visibility_of_element_located((By.XPATH, "//textarea[@autoidentifier='txtbox_Resolution']"))
    ) 
    subir_aba()
    conferencia_texto_resolucao = resolucao.get_attribute("value")
    conferencia_status_chamado = testar_status()
    tickets_concluidos = openpyxl.load_workbook('tickets_concluidos.xlsx')
    pagina_tickets = tickets_concluidos['Plan1']
    pagina_tickets.append([chamado,"Concluido", conferencia_texto_resolucao, conferencia_status_chamado.text, motivo, descrição, Abertoem, Atualizado, Prioridade, Solicitante, Organizaçãodosolicitante, Atribuído, Grupoatribuído, Tipodeticket, Resolverem, StatusdoSLA, Organizaçãodobeneficiário, SelecioneaUnidade, Categorização])
    tickets_concluidos.save('tickets_concluidos.xlsx')

def subir_aba():
    driver.execute_script("window.scrollTo(0, 0);")

logar()

for linha in pagina_lista_chamados.iter_rows(min_row=676, values_only=True):
    chamado, status, motivo, descrição, Abertoem, Atualizado, Prioridade, Solicitante, Organizaçãodosolicitante, Atribuído, Grupoatribuído, Tipodeticket, Resolverem, StatusdoSLA, Organizaçãodobeneficiário, SelecioneaUnidade, Categorização = linha

    pesquisar_chamado()
    status_chamado = testar_status()
    
    if status_chamado.text == "Resolvido":
        alimentar_nova_planilha()
        fechar_chamado()
        continue
        
    if status_chamado.text == "Pendente Fornecedor":
        botao_de_acao()
        sleep(0.5)
        botao_de_retomar()
        atender_chamado()
    
    elif status_chamado.text == "Cadastro contábil":
        resolver_chamado()

    elif status_chamado.text == "Aceitação":
        botao_de_acao()
        botao_de_aceitacao()
        atender_chamado()

    elif status_chamado.text == "Pendente Material":
        botao_de_acao()
        sleep(0.5)
        botao_de_retomar()
        atender_chamado()

    elif status_chamado.text == "Resposta do Cliente":
        botao_de_acao()
        sleep(0.5)
        botao_de_aceitacao()
        atender_chamado()
    
    elif status_chamado.text == "Execução pendente":
        botao_de_acao()
        sleep(0.5)
        botao_de_retomar()
        atender_chamado()

    elif status_chamado.text == "Em andamento":
        atender_chamado()

    else:
        atender_chamado()
